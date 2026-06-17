import hashlib
import json
import re
import shutil
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path

from docx import Document
from openpyxl import Workbook, load_workbook
from PyPDF2 import PdfReader


ROOT = Path(r"C:\Users\HP 250 G10\Documents\GITHUT\MEC")
SOURCE_ROOTS = [ROOT / "FUENTES_ORIGINALES" / "MEC", ROOT / "FUENTES_ORIGINALES" / "MEC2"]
DEST_ROOT = ROOT / "INSTITUCIONES"

INSTITUTIONS = {
    "COLEGIO_MANUEL_MOLINAS": [
        "MANUEL MOLINAS",
        "MANUAL MOLINA",
        "MANUEL MOLINA",
    ],
    "DEFENSORES_DEL_CHACO": [
        "DEFENSORES DEL CHACO",
        "VH ESC DEFENSORES",
        "VH COL. NAC. DEFENSORES",
    ],
    "CENTRO_EDUCATIVO_MARANGATU": [
        "MARANGATU",
    ],
    "COLEGIO_NACIONAL_DON_JORGE_GAYOSO": [
        "JORGE GAYOSO",
        "DON JORGE GAYOSO",
    ],
}

SKIP_PATH_PARTS = {
    "SEGUIMIENTO-ESCOLAR",
    "A ELIMINAR",
    "TEMP-REPO",
    "NODE_MODULES",
    ".GIT",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def extract_text_for_institution(path: Path) -> str:
    suffix = path.suffix.lower()
    try:
        if suffix in {".txt", ".md", ".html", ".htm", ".json", ".xml", ".js", ".ts", ".tsx", ".css", ".ps1"}:
            return path.read_text(encoding="utf-8", errors="ignore")[:12000]
        if suffix == ".docx":
            doc = Document(str(path))
            return "\n".join(paragraph.text for paragraph in doc.paragraphs[:80])[:12000]
        if suffix == ".xlsx":
            workbook = load_workbook(path, data_only=True, read_only=True)
            parts = []
            for sheet in workbook.worksheets[:3]:
                parts.append(sheet.title)
                for row in sheet.iter_rows(min_row=1, max_row=30, max_col=8, values_only=True):
                    values = ["" if value is None else str(value) for value in row]
                    parts.append(" | ".join(values))
            return "\n".join(parts)[:12000]
        if suffix == ".pdf":
            reader = PdfReader(str(path))
            parts = []
            for page in reader.pages[:3]:
                parts.append(page.extract_text() or "")
            return "\n".join(parts)[:12000]
    except Exception:
        return ""
    return ""


def slug(text: str) -> str:
    text = re.sub(r"\s+", "_", text.strip())
    text = re.sub(r'[<>:"/\\|?*]+', "_", text)
    return text


def detect_institution(path: Path):
    path_upper = str(path).upper()
    file_upper = path.name.upper()
    parent_parts = [part.upper() for part in path.parts[-5:]]
    content_upper = extract_text_for_institution(path).upper()

    scores = Counter()
    reasons = defaultdict(list)

    for institution, aliases in INSTITUTIONS.items():
        for alias in aliases:
            alias_upper = alias.upper()
            if alias_upper in file_upper:
                scores[institution] += 6
                reasons[institution].append(f"filename:{alias}")
            if any(alias_upper in part for part in parent_parts):
                scores[institution] += 3
                reasons[institution].append(f"path:{alias}")
            if alias_upper in path_upper:
                scores[institution] += 1
            if alias_upper and alias_upper in content_upper:
                scores[institution] += 10
                reasons[institution].append(f"content:{alias}")

    if not scores:
        return None, []

    best_institution, best_score = scores.most_common(1)[0]
    if best_score <= 0:
        return None, []
    return best_institution, reasons[best_institution]


def should_skip(path: Path) -> bool:
    upper_parts = {part.upper() for part in path.parts}
    return any(part in upper_parts for part in SKIP_PATH_PARTS)


def detect_grade(text: str) -> str:
    upper = text.upper()
    if re.search(r"7_GRADO|\b7(MO|NO)?\b|7º|SEPTIMO|S[ÉE]PTIMO", upper):
        return "7_Grado"
    if re.search(r"8_GRADO|\b8(VO|MO|NO)?\b|8º|OCTAVO", upper):
        return "8_Grado"
    if re.search(r"9_GRADO|\b9(NO|MO)?\b|9º|NOVENO", upper):
        return "9_Grado"
    if re.search(r"\b1(RO)?\b|1º|PRIMER|1_CURSO", upper):
        return "1_Curso"
    if re.search(r"\b2(DO)?\b|2º|SEGUNDO|2_CURSO", upper):
        return "2_Curso"
    if re.search(r"\b3(RO)?\b|3º|TERCER|3_CURSO", upper):
        return "3_Curso"
    return "Sin_Grado"


def detect_subject(text: str) -> str:
    upper = text.upper()
    mapping = [
        ("Etica", ["ETICA", "ÉTICA", "FORMACION ETICA"]),
        ("Ciencias", ["CIENCIAS", "CIENCIAS NATURALES"]),
        ("Resistencia", ["RESISTENCIA"]),
        ("Tecnicas_Instrumentales", ["TECNICAS INSTRUMENTALES", "TÉCNICAS INSTRUMENTALES"]),
        ("Orientacion", ["ORIENTACION", "ORIENTACIÓN"]),
        ("Educacion_Ambiental", ["EDUCACION AMBIENTAL", "EDUCACIÓN AMBIENTAL"]),
        ("Metodologia", ["METODOLOGIA", "METODOLOGÍA"]),
        ("Sociologia_Urbana_y_Rural", ["SOCIOLOGIA URBANA Y RURAL", "SOCIOLOGÍA URBANA Y RURAL"]),
        ("Diagnostico", ["DIAGNOSTICO", "DIAGNÓSTICO"]),
        ("Institucional", ["INSTITUCIONAL", "PEI", "PCI", "POA"]),
    ]
    for subject, aliases in mapping:
        if any(alias in upper for alias in aliases):
            return subject
    return "Sin_Asignatura"


def detect_turn(text: str) -> str:
    upper = text.upper()
    if re.search(r"\bTM\b", upper):
        return "Turno_TM"
    if re.search(r"\bTT\b", upper):
        return "Turno_TT"
    return "General"


def collect_source_files():
    files = []
    for root in SOURCE_ROOTS:
        for path in root.rglob("*"):
            if not path.is_file():
                continue
            if should_skip(path):
                continue
            institution, reasons = detect_institution(path)
            if institution:
                files.append((institution, path, reasons))
    return sorted(files, key=lambda item: str(item[1]).lower())


def build_inventory(items):
    inventory = []
    for institution, path, reasons in items:
        text = str(path)
        inventory.append(
            {
                "institution": institution,
                "source": str(path),
                "name": path.name,
                "suffix": path.suffix.lower(),
                "size": path.stat().st_size,
                "sha256": sha256_file(path),
                "grade": detect_grade(text),
                "subject": detect_subject(text),
                "turn": detect_turn(text),
                "institution_reasons": reasons,
            }
        )
    return inventory


def enrich_metadata_from_hash_groups(inventory):
    by_hash = defaultdict(list)
    for item in inventory:
        by_hash[(item["institution"], item["sha256"])].append(item)

    for items in by_hash.values():
        known_grades = [item["grade"] for item in items if item["grade"] != "Sin_Grado"]
        known_subjects = [item["subject"] for item in items if item["subject"] != "Sin_Asignatura"]
        known_turns = [item["turn"] for item in items if item["turn"] != "General"]

        inferred_grade = known_grades[0] if known_grades else None
        inferred_subject = known_subjects[0] if known_subjects else None
        inferred_turn = known_turns[0] if len(set(known_turns)) == 1 and known_turns else None

        for item in items:
            if inferred_grade and item["grade"] == "Sin_Grado":
                item["grade"] = inferred_grade
            if inferred_subject and item["subject"] == "Sin_Asignatura":
                item["subject"] = inferred_subject
            if inferred_turn and item["turn"] == "General":
                item["turn"] = inferred_turn

    return inventory


def choose_bucket(items):
    turns = {item["turn"] for item in items}
    if len(turns) > 1:
        return "General"
    return items[0]["turn"]


def copy_unique_file(source: Path, target_dir: Path, file_name: str, file_hash: str):
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / file_name
    if not target.exists():
        shutil.copy2(source, target)
        return str(target)
    if sha256_file(target) == file_hash:
        return str(target)
    conflict_target = target_dir / f"{target.stem}__TMP_{file_hash[:8]}{target.suffix}"
    if not conflict_target.exists():
        shutil.copy2(source, conflict_target)
    return str(conflict_target)


def extract_workbook_preview(path: Path):
    workbook = load_workbook(path, data_only=False)
    preview = []
    for sheet in workbook.worksheets:
        preview.append(f"### Hoja: {sheet.title}")
        max_row = min(sheet.max_row, 15)
        max_col = min(sheet.max_column, 8)
        for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            cells = ["" if value is None else str(value) for value in row]
            preview.append(" | ".join(cells).rstrip())
        preview.append("")
    return "\n".join(preview).strip()


def merge_workbooks(paths, target_path: Path):
    merged = Workbook()
    merged.remove(merged.active)
    for index, path in enumerate(paths, start=1):
        workbook = load_workbook(path, data_only=False)
        prefix = f"V{index}"
        for sheet in workbook.worksheets:
            title = slug(f"{prefix}_{sheet.title}")[:31]
            new_sheet = merged.create_sheet(title)
            for row in sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.border = copy(cell.border)
                        new_cell.alignment = copy(cell.alignment)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                    if cell.hyperlink:
                        new_cell._hyperlink = copy(cell.hyperlink)
                    if cell.comment:
                        new_cell.comment = copy(cell.comment)
            for column_letter, dim in sheet.column_dimensions.items():
                new_sheet.column_dimensions[column_letter].width = dim.width
            for row_idx, dim in sheet.row_dimensions.items():
                new_sheet.row_dimensions[row_idx].height = dim.height
            for merged_range in sheet.merged_cells.ranges:
                new_sheet.merge_cells(str(merged_range))
    target_path.parent.mkdir(parents=True, exist_ok=True)
    merged.save(target_path)


def write_conflict_summary(paths, target_dir: Path, base_name: str):
    summary_name = f"{base_name}__COMBINADO.md"
    summary_path = target_dir / summary_name
    lines = [
        f"# Consolidacion de {base_name}",
        "",
        "Este archivo resume varias versiones con el mismo nombre y contenido distinto.",
        "",
    ]
    for index, path in enumerate(paths, start=1):
        lines.append(f"## Version {index}")
        lines.append(f"Fuente: `{path}`")
        lines.append("")
        if path.suffix.lower().endswith("xlsx"):
            lines.append(extract_workbook_preview(path))
        else:
            lines.append("Vista previa no automatizada para este formato.")
        lines.append("")
    summary_path.write_text("\n".join(lines), encoding="utf-8")
    return str(summary_path)


def process_institution(institution, items):
    institution_root = DEST_ROOT / institution
    if institution_root.exists():
        shutil.rmtree(institution_root)
    institution_root.mkdir(parents=True, exist_ok=True)

    by_hash = defaultdict(list)
    for item in items:
        by_hash[item["sha256"]].append(item)

    copied = []
    deduplicated = []
    for _, group in by_hash.items():
        representative = group[0]
        bucket = choose_bucket(group)
        target_dir = institution_root / representative["grade"] / representative["subject"] / bucket
        target_path = copy_unique_file(
            Path(representative["source"]),
            target_dir,
            representative["name"],
            representative["sha256"],
        )
        copied.append(
            {
                "target": target_path,
                "original_name": representative["name"],
                "sources": [item["source"] for item in group],
                "sha256": representative["sha256"],
            }
        )
        if len(group) > 1:
            deduplicated.append(
                {
                    "kept_target": target_path,
                    "duplicate_sources": [item["source"] for item in group[1:]],
                }
            )

    name_groups = defaultdict(list)
    for entry in copied:
        key = (str(Path(entry["target"]).parent), entry["original_name"].lower())
        name_groups[key].append(entry)

    conflicts = []
    for (target_dir_str, file_name), entries in name_groups.items():
        if len(entries) < 2:
            continue
        hashes = {entry["sha256"] for entry in entries}
        if len(hashes) == 1:
            continue

        target_dir = Path(target_dir_str)
        stem = Path(file_name).stem
        suffix = Path(file_name).suffix
        source_paths = [Path(entry["sources"][0]) for entry in entries]

        for index, entry in enumerate(entries, start=1):
            current_target = Path(entry["target"])
            renamed = current_target.with_name(f"{stem}__V{index}{suffix}")
            if current_target.exists() and not renamed.exists():
                current_target.rename(renamed)
            entry["target"] = str(renamed)

        summary_path = write_conflict_summary(source_paths, target_dir, stem)
        combined_workbook = None
        if suffix.lower().endswith("xlsx"):
            combined_workbook_path = target_dir / f"{stem}__COMBINADO{suffix}"
            merge_workbooks(source_paths, combined_workbook_path)
            combined_workbook = str(combined_workbook_path)

        conflicts.append(
            {
                "folder": str(target_dir),
                "name": file_name,
                "versions": [entry["target"] for entry in entries],
                "sources": [str(path) for path in source_paths],
                "combined_summary": summary_path,
                "combined_workbook": combined_workbook,
            }
        )

    report = {
        "institution": institution,
        "source_file_count": len(items),
        "unique_content_count": len(by_hash),
        "copied_entries": copied,
        "deduplicated_content": deduplicated,
        "same_name_different_content": conflicts,
    }

    (institution_root / "00_REPORTE_ORGANIZACION.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    lines = [
        "# Resumen de organización",
        "",
        f"- Institución: `{institution}`",
        f"- Archivos fuente detectados: `{len(items)}`",
        f"- Contenidos únicos copiados: `{len(by_hash)}`",
        f"- Duplicados idénticos consolidados: `{len(deduplicated)}`",
        f"- Conflictos de mismo nombre con contenido distinto: `{len(conflicts)}`",
        "",
        "## Estructura aplicada",
        "",
        "- Organización por `institución/grado/asignatura/turno`.",
        "- Los archivos idénticos se copiaron una sola vez.",
        "- Cuando hubo mismo nombre y contenido distinto, se conservaron versiones separadas y se creó un archivo `__COMBINADO`.",
    ]
    (institution_root / "00_RESUMEN_ORGANIZACION.md").write_text("\n".join(lines), encoding="utf-8")
    return report


def main():
    items = collect_source_files()
    inventory = enrich_metadata_from_hash_groups(build_inventory(items))

    by_institution = defaultdict(list)
    for item in inventory:
        by_institution[item["institution"]].append(item)

    reports = {}
    global_summary = []
    audit_rows = []
    for institution, institution_items in sorted(by_institution.items()):
        report = process_institution(institution, institution_items)
        reports[institution] = report
        global_summary.append(
            {
                "institution": institution,
                "source_file_count": report["source_file_count"],
                "unique_content_count": report["unique_content_count"],
                "deduplicated_count": len(report["deduplicated_content"]),
                "conflict_count": len(report["same_name_different_content"]),
            }
        )
        for item in institution_items:
            audit_rows.append(
                {
                    "institution": institution,
                    "source": item["source"],
                    "grade": item["grade"],
                    "subject": item["subject"],
                    "turn": item["turn"],
                    "reasons": item.get("institution_reasons", []),
                }
            )

    (DEST_ROOT / "00_REPORTE_GLOBAL_INSTITUCIONES.json").write_text(
        json.dumps(global_summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (DEST_ROOT / "00_AUDITORIA_IDENTIFICACION_INSTITUCION.json").write_text(
        json.dumps(audit_rows, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
