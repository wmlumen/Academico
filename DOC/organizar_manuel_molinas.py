import hashlib
import json
import os
import re
import shutil
from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook


ROOT = Path(r"C:\Users\HP 250 G10\Documents\GITHUT\MEC")
SOURCE_ROOT = ROOT / "FUENTES_ORIGINALES" / "MEC2"
SOURCE_ROOT_2 = ROOT / "FUENTES_ORIGINALES" / "MEC"
EEB_ROOT = ROOT / "CURRICULO_BASE" / "EEB"
DEST_ROOT = ROOT / "INSTITUCIONES" / "COLEGIO_MANUEL_MOLINAS"
REPORT_PATH = DEST_ROOT / "00_REPORTE_ORGANIZACION.json"
SUMMARY_PATH = DEST_ROOT / "00_RESUMEN_ORGANIZACION.md"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def slug(text: str) -> str:
    text = re.sub(r"\s+", "_", text.strip())
    text = re.sub(r'[<>:"/\\|?*]+', "_", text)
    return text


def detect_grade(text: str) -> str:
    upper = text.upper()
    if re.search(r"7_GRADO|\b7(MO|NO)?\b|7º|SEPTIMO|S[ÉE]PTIMO", upper):
        return "7_Grado"
    if re.search(r"8_GRADO|\b8(MO|NO)?\b|8º|OCTAVO", upper):
        return "8_Grado"
    if re.search(r"9_GRADO|\b9(MO|NO)?\b|9º|NOVENO", upper):
        return "9_Grado"
    return "Sin_Grado"


def detect_subject(text: str) -> str:
    upper = text.upper()
    if "ETICA" in upper or "ÉTICA" in text.upper():
        return "Etica"
    if "CIENCIA" in upper:
        return "Ciencias"
    return "Sin_Asignatura"


def detect_turn(text: str) -> str:
    upper = text.upper()
    if re.search(r"\bTM\b", upper):
        return "Turno_TM"
    if re.search(r"\bTT\b", upper):
        return "Turno_TT"
    return "General"


def collect_source_files():
    candidates = []
    for source_base in (SOURCE_ROOT, SOURCE_ROOT_2):
        for path in source_base.rglob("*"):
            if not path.is_file():
                continue
            upper = str(path).upper()
            if any(token in upper for token in ("MANUEL", "MOLINA", "MOLINAS")):
                candidates.append(path)

    for path in EEB_ROOT.rglob("*"):
        if not path.is_file():
            continue
        upper = str(path).upper()
        if any(token in upper for token in ("7_GRADO", "8_GRADO", "9_GRADO")):
            candidates.append(path)
    return sorted(candidates)


def build_inventory(paths):
    inventory = []
    for path in paths:
        text = str(path)
        inventory.append(
            {
                "source": str(path),
                "name": path.name,
                "suffix": path.suffix.lower(),
                "size": path.stat().st_size,
                "sha256": sha256_file(path),
                "grade": detect_grade(text),
                "subject": detect_subject(text),
                "turn": detect_turn(text),
            }
        )
    return inventory


def enrich_metadata_from_hash_groups(inventory):
    by_hash = defaultdict(list)
    for item in inventory:
        by_hash[item["sha256"]].append(item)

    for items in by_hash.values():
        known_grades = [item["grade"] for item in items if item["grade"] != "Sin_Grado"]
        known_subjects = [item["subject"] for item in items if item["subject"] != "Sin_Asignatura"]
        known_turns = [item["turn"] for item in items if item["turn"] != "General"]

        inferred_grade = known_grades[0] if known_grades else None
        inferred_subject = known_subjects[0] if known_subjects else None
        inferred_turn = known_turns[0] if len(set(known_turns)) == 1 and known_turns else None

        for item in items:
            if inferred_grade and item["grade"] == "Sin_Grado":
                item["grade"] = inferred_grade
            if inferred_subject and item["subject"] == "Sin_Asignatura":
                item["subject"] = inferred_subject
            if inferred_turn and item["turn"] == "General":
                item["turn"] = inferred_turn

    return inventory


def choose_bucket(items):
    turns = {item["turn"] for item in items}
    if len(turns) > 1:
        return "General"
    return items[0]["turn"]


def copy_unique_file(source: Path, target_dir: Path, file_name: str, file_hash: str):
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / file_name
    if not target.exists():
        shutil.copy2(source, target)
        return str(target)
    if sha256_file(target) == file_hash:
        return str(target)
    conflict_target = target_dir / f"{target.stem}__TMP_{file_hash[:8]}{target.suffix}"
    if not conflict_target.exists():
        shutil.copy2(source, conflict_target)
    return str(conflict_target)


def extract_workbook_preview(path: Path):
    workbook = load_workbook(path, data_only=False)
    preview = []
    for sheet in workbook.worksheets:
        preview.append(f"### Hoja: {sheet.title}")
        max_row = min(sheet.max_row, 15)
        max_col = min(sheet.max_column, 8)
        for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
            cells = ["" if value is None else str(value) for value in row]
            preview.append(" | ".join(cells).rstrip())
        preview.append("")
    return "\n".join(preview).strip()


def merge_workbooks(paths, target_path: Path):
    merged = Workbook()
    merged.remove(merged.active)
    for index, path in enumerate(paths, start=1):
        workbook = load_workbook(path, data_only=False)
        prefix = f"V{index}"
        for sheet in workbook.worksheets:
            title = slug(f"{prefix}_{sheet.title}")[:31]
            new_sheet = merged.create_sheet(title)
            for row in sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.fill = copy(cell.fill)
                        new_cell.border = copy(cell.border)
                        new_cell.alignment = copy(cell.alignment)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                    if cell.hyperlink:
                        new_cell._hyperlink = copy(cell.hyperlink)
                    if cell.comment:
                        new_cell.comment = copy(cell.comment)
            for column_letter, dim in sheet.column_dimensions.items():
                new_sheet.column_dimensions[column_letter].width = dim.width
            for row_idx, dim in sheet.row_dimensions.items():
                new_sheet.row_dimensions[row_idx].height = dim.height
            for merged_range in sheet.merged_cells.ranges:
                new_sheet.merge_cells(str(merged_range))
    target_path.parent.mkdir(parents=True, exist_ok=True)
    merged.save(target_path)


def write_conflict_summary(paths, target_dir: Path, base_name: str):
    summary_name = f"{base_name}__COMBINADO.md"
    summary_path = target_dir / summary_name
    lines = [
        f"# Consolidacion de {base_name}",
        "",
        "Este archivo resume varias versiones con el mismo nombre y contenido distinto.",
        "",
    ]
    for index, path in enumerate(paths, start=1):
        lines.append(f"## Version {index}")
        lines.append(f"Fuente: `{path}`")
        lines.append("")
        if path.suffix.lower().endswith("xlsx"):
            lines.append(extract_workbook_preview(path))
        else:
            lines.append("Vista previa no automatizada para este formato.")
        lines.append("")
    summary_path.write_text("\n".join(lines), encoding="utf-8")
    return str(summary_path)


def main():
    source_files = collect_source_files()
    inventory = build_inventory(source_files)
    inventory = enrich_metadata_from_hash_groups(inventory)

    if DEST_ROOT.exists():
        shutil.rmtree(DEST_ROOT)
    DEST_ROOT.mkdir(parents=True, exist_ok=True)

    by_hash = defaultdict(list)
    for item in inventory:
        by_hash[item["sha256"]].append(item)

    copied = []
    deduplicated = []
    for _, items in by_hash.items():
        representative = items[0]
        bucket = choose_bucket(items)
        target_dir = DEST_ROOT / representative["grade"] / representative["subject"] / bucket
        target_path = copy_unique_file(
            Path(representative["source"]),
            target_dir,
            representative["name"],
            representative["sha256"],
        )
        copied.append(
            {
                "target": target_path,
                "original_name": representative["name"],
                "sources": [item["source"] for item in items],
                "sha256": representative["sha256"],
            }
        )
        if len(items) > 1:
            deduplicated.append(
                {
                    "kept_target": target_path,
                    "duplicate_sources": [item["source"] for item in items[1:]],
                }
            )

    name_groups = defaultdict(list)
    for entry in copied:
        target = Path(entry["target"])
        key = (str(target.parent), entry["original_name"].lower())
        name_groups[key].append(entry)

    conflicts = []
    for (target_dir_str, file_name), entries in name_groups.items():
        if len(entries) < 2:
            continue
        hashes = {entry["sha256"] for entry in entries}
        if len(hashes) == 1:
            continue
        target_dir = Path(target_dir_str)
        stem = Path(file_name).stem
        suffix = Path(file_name).suffix
        source_paths = [Path(entry["sources"][0]) for entry in entries]

        for index, entry in enumerate(entries, start=1):
            current_target = Path(entry["target"])
            renamed = current_target.with_name(f"{stem}__V{index}{suffix}")
            if current_target.exists() and not renamed.exists():
                current_target.rename(renamed)
            entry["target"] = str(renamed)

        summary_path = write_conflict_summary(source_paths, target_dir, stem)
        combined_workbook = None
        if suffix.lower().endswith("xlsx"):
            combined_workbook_path = target_dir / f"{stem}__COMBINADO{suffix}"
            merge_workbooks(source_paths, combined_workbook_path)
            combined_workbook = str(combined_workbook_path)

        conflicts.append(
            {
                "folder": str(target_dir),
                "name": file_name,
                "versions": [entry["target"] for entry in entries],
                "sources": [str(path) for path in source_paths],
                "combined_summary": summary_path,
                "combined_workbook": combined_workbook,
            }
        )

    report = {
        "source_file_count": len(source_files),
        "unique_content_count": len(by_hash),
        "copied_entries": copied,
        "deduplicated_content": deduplicated,
        "same_name_different_content": conflicts,
    }

    REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    summary_lines = [
        "# Resumen de organización",
        "",
        f"- Institución: `COLEGIO_MANUEL_MOLINAS`",
        f"- Archivos fuente detectados: `{len(source_files)}`",
        f"- Contenidos únicos copiados: `{len(by_hash)}`",
        f"- Duplicados idénticos consolidados: `{len(deduplicated)}`",
        f"- Conflictos de mismo nombre con contenido distinto: `{len(conflicts)}`",
        "",
        "## Estructura aplicada",
        "",
        "- Organización por `institución/grado/asignatura/turno`.",
        "- Los archivos idénticos se copiaron una sola vez.",
        "- Cuando hubo mismo nombre y contenido distinto, se conservaron versiones separadas y se creó un archivo `__COMBINADO`.",
        "",
        "## Conflictos resueltos",
        "",
    ]

    if conflicts:
        for conflict in conflicts:
            summary_lines.append(f"- `{conflict['name']}` en `{conflict['folder']}`")
            summary_lines.append(f"  Resumen combinado: `{conflict['combined_summary']}`")
            if conflict["combined_workbook"]:
                summary_lines.append(f"  Archivo combinado: `{conflict['combined_workbook']}`")
    else:
        summary_lines.append("- No se detectaron conflictos de mismo nombre con contenido distinto.")

    SUMMARY_PATH.write_text("\n".join(summary_lines), encoding="utf-8")


if __name__ == "__main__":
    main()
